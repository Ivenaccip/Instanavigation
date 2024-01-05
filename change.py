import pandas as pd
from openpyxl import Workbook
from sqlalchemy import create_engine
import mysql.connector
import subprocess
import sys

#Hashtags
def hashtags(cursor):
    cursor.execute("SELECT `company`, `identificators` FROM companies")
    df_hashtags = pd.DataFrame(cursor.fetchall(), columns=['company', 'identificators'])
    cursor.close()
    return df_hashtags

#Id objects
def id_obj_down(cursor):
    query_id_obj = "SELECT Object_ID, Type, Link, Extract_text, Fecha, Profile FROM id_obj_download"
    cursor.execute(query_id_obj, create_engine)
    db_data = cursor.fetchall()
    cursor.close()
    return db_data

def reward(cursor, link):
    reward = cursor.execute(f"SELECT reach FROM reach_profile WHERE name = %s", (link))
    return reward

# Variables iniciales
information_xlsx = 'Information.xlsx'

conn = mysql.connector.connect(
#Conexion SQL
    host = 'localhost',
    user = 'office',
    password = 'Kroon111',
    database = 'alpha_test'
)
# Usando las funciones
cursor = conn.cursor()

df_hashtags = hashtags(cursor)
db_data = id_obj_down(cursor)
brand_results = {}

for marca, hashtags_brand in df_hashtags.iterrows():
    find_results = []

    for record in db_data:
        obj_id, tipo_obj, link, texto, date, instagram  = record
        print(f"It's {tipo_obj}")
        matched_brands = []  # Lista para almacenar las marcas que coinciden

        if tipo_obj == "Post":
            for hashtag in hashtags_brand.dropna():
                if hashtag in texto:
                    matched_brands.append(marca)
                    subprocess.run(["./reach.sh", link])
                    subprocess.run(["python3", "check.py", link])
                    reward = reward(cursor, link)

            # Verifica si hay coincidencias múltiples y actualiza brand_results
            if matched_brands:
                result = (texto, date, link, reward)
                for brand in matched_brands:
                    if brand not in brand_results:
                        brand_results[brand] = []
                    brand_results[brand].append(result)

                if len(matched_brands) >= 3:
                    general_key = 'General'
                    if general_key not in brand_results:
                        brand_results[general_key] = []
                    brand_results[general_key].append(result)
        else:
            subprocess.run(["./tesserach.sh", obj_id])

# Crear un archivo Excel de salida y hojas de cálculo con nombres de archivo y marca
output_filename = 'Results.xlsx'
workbook = Workbook()
if len(sys.argv) > 1:
    name = sys.argv[1]
else:
    print("There is a problem with argument line command")

for marca, resultados in brand_results.items():
    worksheet = workbook.create_sheet(title=f'{marca}')
    worksheet.append(["Naam", "Datum", "Medium", "url/ID", "Reach", "Media"])  # Agregar la primera fila con encabezados

    for idx, (texto, date, link, reward) in enumerate(resultados, start=2):  # Comenzar desde la fila 2 para los datos
        worksheet.cell(row=idx, column=1, value=name)
        worksheet.cell(row=idx, column=2, value=date)
        worksheet.cell(row=idx, column=3, value=tipo_obj)
        worksheet.cell(row=idx, column=4, value=link)
        worksheet.cell(row=idx, column=5, value=reward)
        worksheet.cell(row=idx, column=6, value=link)

# Guardar el archivo Excel
workbook.remove(workbook['Sheet'])  # Eliminar la hoja en blanco predeterminada
workbook.save(output_filename)
#name, datum medium,url,reach, media value
